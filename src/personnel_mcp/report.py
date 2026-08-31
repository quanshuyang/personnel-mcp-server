"""Excel report builder for personnel statistics.

Reuses the normalization and analytics implemented in `data` and `analytics`
so the exported report stays consistent with the MCP tool outputs.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .analytics import (
    department_statistics,
    filter_employees,
    headcount_summary,
    validate_personnel_data,
)
from .data import PersonnelDataError, PersonnelDataset, get_data_dir, normalize_text

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=12)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_table(sheet, headers: list[str], rows: list[list[Any]], start_row: int = 1) -> int:
    """Write a header row then data rows, returning the next free row index."""
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    for row_offset, row_values in enumerate(rows, start=start_row + 1):
        for column, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_offset, column=column, value=value)
            cell.alignment = CENTER if not isinstance(value, str) or len(value) <= 20 else LEFT
    return start_row + len(rows) + 2


def _autosize(sheet, max_width: int = 48) -> None:
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = max(len(str(cell.value).encode("gbk", "ignore")) // 2, len(str(cell.value)))
            widths[cell.column] = min(max(widths.get(cell.column, 0), length + 2), max_width)
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def build_report(dataset: PersonnelDataset, known_departments: list[str] | None = None) -> dict[str, Any]:
    """Assemble all statistics and return the workbook plus a machine-readable summary."""
    summary = headcount_summary(dataset)
    dept_stats = department_statistics(dataset)
    validation = validate_personnel_data(dataset, known_departments)
    normalized_rows = filter_employees(dataset, limit=len(dataset.records))["employees"]

    workbook = Workbook()
    overview = workbook.active
    overview.title = "汇总"

    overview.cell(row=1, column=1, value="人员统计汇总").font = TITLE_FONT
    overview.cell(row=2, column=1, value=f"来源文件: {dataset.source_file}")
    if dataset.sheet_name:
        overview.cell(row=3, column=1, value=f"工作表: {dataset.sheet_name}")

    gender = summary["gender"]
    overview.cell(row=5, column=1, value="指标")
    overview.cell(row=5, column=2, value="数值")
    for cell in (overview.cell(row=5, column=1), overview.cell(row=5, column=2)):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    overview_rows = [
        ["总人数", summary["total_headcount"]],
        ["部门数", summary["department_count"]],
        ["男性人数", gender["male"]],
        ["女性人数", gender["female"]],
        ["未识别性别人数", gender["unknown"]],
        ["男性比例", gender["male_ratio"]],
        ["女性比例", gender["female_ratio"]],
        ["数据问题数", validation["summary"]["total_issues"]],
    ]
    for offset, (label, value) in enumerate(overview_rows, start=6):
        overview.cell(row=offset, column=1, value=label).alignment = LEFT
        overview.cell(row=offset, column=2, value=value).alignment = CENTER

    next_row = _write_table(
        overview,
        ["表头映射"],
        [[field, header] for field, header in dataset.mapped_columns.items()],
        start_row=len(overview_rows) + 8,
    )
    overview.cell(row=next_row - 2, column=1, value="字段映射").font = TITLE_FONT

    # 部门统计
    dept_sheet = workbook.create_sheet("部门统计")
    dept_rows = []
    for item in dept_stats["departments"]:
        g = item["gender"]
        dept_rows.append([item["department"], item["headcount"], g["male"], g["female"], g["unknown"]])
    if dept_stats["unassigned_headcount"]:
        dept_rows.append(["（未分配部门）", dept_stats["unassigned_headcount"], "", "", ""])
    _write_table(dept_sheet, ["部门", "人数", "男性", "女性", "未识别性别"], dept_rows)

    # 性别统计
    gender_sheet = workbook.create_sheet("性别统计")
    gender_rows = [
        ["男", gender["male"], gender["male_ratio"]],
        ["女", gender["female"], gender["female_ratio"]],
        ["未填写/其他", gender["unknown"], ""],
    ]
    _write_table(gender_sheet, ["性别", "人数", "比例"], gender_rows)

    # 数据问题
    issue_sheet = workbook.create_sheet("数据问题")
    issue_rows = [[issue["severity"], issue["code"], issue["field"], ", ".join(map(str, issue["rows"])), issue["message"]] for issue in validation["issues"]]
    if not issue_rows:
        issue_rows = [["-", "-", "-", "-", "未发现数据问题"]]
    _write_table(issue_sheet, ["级别", "代码", "字段", "行号", "说明"], issue_rows)

    # 规范化数据
    norm_sheet = workbook.create_sheet("规范化数据")
    norm_rows = [
        [record.get("employee_id", ""), record.get("name", ""), record.get("department", ""), record.get("gender", ""), record.get("gender_normalized", "unknown"), record.get("source_row")]
        for record in normalized_rows
    ]
    _write_table(norm_sheet, ["工号", "姓名", "部门", "原始性别", "规范化性别", "来源行"], norm_rows)

    for sheet in workbook.worksheets:
        _autosize(sheet)

    return {
        "workbook": workbook,
        "summary": summary,
        "department_statistics": dept_stats,
        "validation": validation,
        "normalized_records": normalized_rows,
    }


def resolve_output_path(output_file: str, data_dir: Path | None = None) -> Path:
    """Resolve an output path that must stay inside the data directory."""
    if not output_file or not output_file.strip():
        raise PersonnelDataError("output_file 不能为空。")

    candidate = Path(output_file)
    if candidate.is_absolute():
        raise PersonnelDataError("output_file 必须是数据目录内的相对路径。")

    base_dir = (data_dir or get_data_dir()).resolve()
    path = (base_dir / candidate).resolve()
    try:
        path.relative_to(base_dir)
    except ValueError as exc:
        raise PersonnelDataError("output_file 不能指向数据目录外的文件。") from exc

    if path.suffix.lower() != ".xlsx":
        raise PersonnelDataError("报告仅支持导出为 .xlsx 文件。")
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def export_report(
    dataset: PersonnelDataset,
    output_file: str,
    known_departments: list[str] | None = None,
    data_dir: Path | None = None,
) -> dict[str, Any]:
    """Build the statistics workbook and write it to a data-directory-relative path."""
    built = build_report(dataset, known_departments)
    output_path = resolve_output_path(output_file, data_dir)
    built["workbook"].save(output_path)
    return {
        "source": built["summary"]["source"],
        "output_file": str(output_path),
        "total_headcount": built["summary"]["total_headcount"],
        "department_count": built["summary"]["department_count"],
        "issue_count": built["validation"]["summary"]["total_issues"],
        "warnings": dataset.warnings,
    }
